"""Report generator — ported from Resital Villa Scripts utils/reports.py.

Generates all report types:
  - weekly (caretaker-based villa assignments)
  - ismet_abi (reservations with welcome pack + pool heating)
  - hamdi_abi (unassigned villas)
  - korsan (calendar-style Excel workbook)
  - total_history (merged all-time reservations)
"""

import json
import os
import re
import warnings
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from villa_matcher.reports.caretakers import get_caretakers
from villa_matcher.reports.file_utils import (
    list_files_with_extension,
    select_latest_file_with_extension,
)
from villa_matcher.reports.reservations import (
    categorise_by_villas,
    extract_reservations,
    extract_welcome_pack_size,
)


# ── Manual reservation helpers ────────────────────────────────────────────────

def _load_manual_reservations_as_dicts(manual_path: str | None) -> list[dict]:
    """Load manual_reservations.json and convert to the dict format
    expected by categorise_by_villas (Accomodation Name, Holiday Start Date, etc.)."""
    if not manual_path or not os.path.isfile(manual_path):
        return []

    with open(manual_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    result = []
    for entry in raw:
        try:
            start = entry["start"]   # "2026-08-10"
            end = entry["end"]       # "2026-08-17"
        except KeyError:
            continue

        # Convert ISO dates to dd/mm/yy format used by the Excel reports
        try:
            sd = datetime.strptime(start, "%Y-%m-%d")
            ed = datetime.strptime(end, "%Y-%m-%d")
            start_fmt = sd.strftime("%d/%m/%y")
            end_fmt = ed.strftime("%d/%m/%y")
        except ValueError:
            start_fmt = start
            end_fmt = end

        villa_name = entry.get("villa", "")
        # Normalize: always use "Villa X" format to match Excel data
        if villa_name and not villa_name.startswith("Villa "):
            villa_name = f"Villa {villa_name}"

        result.append({
            "Accomodation Name": villa_name,
            "Holiday Start Date": start_fmt,
            "Holiday End Date": end_fmt,
            "Lead Passenger": entry.get("passenger", ""),
            "ExtrasAggregated": entry.get("extras", ""),
            "_manual": True,  # marker to distinguish manual entries
            "_notes": entry.get("notes", ""),
        })

    return result


# ── Helpers ─────────────────────────────────────────────────────────────────

def _filter_extras_text(extras: str) -> str:
    if not isinstance(extras, str):
        return ""
    matches = re.findall(r"(1x Extra - (Welcome Pack\s+\d+-\d+ passengers|Pool Heating)[^,]*)", extras)
    filtered = [m[0] for m in matches]
    return ", ".join(filtered)


def _format_date_columns(df, date_columns: dict):
    """Format date columns in a DataFrame."""
    for col, fmt in date_columns.items():
        if col not in df.columns:
            continue
        if df[col].dtype == "object" and df[col].str.contains(r"\d{2}/\d{2}/\d{2}", na=False).all():
            continue
        try:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime(fmt)
            else:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime(fmt)
        except Exception:
            pass


def _parse_date_dmy(date_str: str) -> datetime | None:
    """Parse a dd/mm/yy date string. Returns None on failure."""
    if not date_str or not isinstance(date_str, str):
        return None
    try:
        return datetime.strptime(date_str.strip(), "%d/%m/%y")
    except ValueError:
        return None


def _is_past_reservation(res: dict, today: date) -> bool:
    """Return True if a reservation has already checked out (end date < today).

    Keeps current stays (check-out today or later) and upcoming stays.
    Reservations with an unparseable end date are kept (not silently dropped).
    """
    end_str = res.get("Holiday End Date", "")
    end = _parse_date_dmy(end_str)
    if end is None:
        return False
    return end.date() < today


def _to_date(val) -> datetime | None:
    """Coerce a value (Timestamp, datetime, or dd/mm/yy string) to a datetime."""
    if val is None:
        return None
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        return _parse_date_dmy(val)
    return None


def _extract_snapshot_date(filename: str) -> date | None:
    """Extract snapshot date from filename (e.g. ..._10-08-2026_unlocked.xlsx)."""
    m = re.search(r"(\d{2})[-_.](\d{2})[-_.](\d{4})", filename)
    if not m:
        return None
    try:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def _sorted_snapshots(folder: str) -> list[str]:
    """List .xlsx snapshot filenames sorted chronologically by embedded date."""
    files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]

    def _key(f: str) -> date:
        d = _extract_snapshot_date(f)
        return d if d is not None else date.min

    return sorted(files, key=_key)


def _find_overlaps(reservations: list[dict]) -> dict:
    """Detect & deduplicate overlapping reservations per villa.

    Returns a dict with:
      - warnings: list of dicts describing each issue found
      - clean_reservations: deduplicated reservation list

    Dedup rule:
      - Manual + Excel with EXACTLY the same dates → same reservation, the
        manual copy is removed SILENTLY (no warning).
      - Manual + Excel with DIFFERENT but overlapping dates → real conflict,
        both kept and warned about.
      - Both Excel or both manual overlapping → warned about.
    """
    # Group by villa name
    by_villa: dict[str, list[dict]] = {}
    for r in reservations:
        name = r.get("Accomodation Name", "")
        if name:
            by_villa.setdefault(name, []).append(r)

    warnings: list[dict] = []
    to_remove: set[int] = set()   # id() of reservations to drop

    for villa, villa_res in by_villa.items():
        # Parse dates for every entry; skip those with unparseable dates
        parsed: list[tuple[int, datetime, datetime, dict]] = []
        for idx, r in enumerate(villa_res):
            start = _parse_date_dmy(
                r.get("Holiday Start Date") or r.get("Holiday Start Day", "")
            )
            end = _parse_date_dmy(r.get("Holiday End Date", ""))
            if start is None or end is None:
                continue
            parsed.append((idx, start, end, r))

        # Check every pair for overlap: [s1, e1) vs [s2, e2)
        for a in range(len(parsed)):
            for b in range(a + 1, len(parsed)):
                _ia, s_a, e_a, r_a = parsed[a]
                _ib, s_b, e_b, r_b = parsed[b]

                if not (s_a < e_b and s_b < e_a):
                    continue  # no overlap

                is_manual_a = r_a.get("_manual", False)
                is_manual_b = r_b.get("_manual", False)

                s_a_str = r_a.get("Holiday Start Date") or r_a.get("Holiday Start Day", "")
                e_a_str = r_a.get("Holiday End Date", "")
                s_b_str = r_b.get("Holiday Start Date") or r_b.get("Holiday Start Day", "")
                e_b_str = r_b.get("Holiday End Date", "")

                if is_manual_a and not is_manual_b:
                    # Manual a vs Excel b
                    if s_a == s_b and e_a == e_b:
                        # Exact same dates → same reservation, drop manual silently
                        to_remove.add(id(r_a))
                    else:
                        # Different overlapping dates → real conflict
                        warnings.append({
                            "villa": villa,
                            "type": "overlap_warning",
                            "range1": f"{s_a_str} - {e_a_str}",
                            "range2": f"{s_b_str} - {e_b_str}",
                        })
                elif is_manual_b and not is_manual_a:
                    # Manual b vs Excel a
                    if s_a == s_b and e_a == e_b:
                        # Exact same dates → same reservation, drop manual silently
                        to_remove.add(id(r_b))
                    else:
                        # Different overlapping dates → real conflict
                        warnings.append({
                            "villa": villa,
                            "type": "overlap_warning",
                            "range1": f"{s_a_str} - {e_a_str}",
                            "range2": f"{s_b_str} - {e_b_str}",
                        })
                else:
                    # Both Excel or both manual → genuine data conflict
                    warnings.append({
                        "villa": villa,
                        "type": "overlap_warning",
                        "range1": f"{s_a_str} - {e_a_str}",
                        "range2": f"{s_b_str} - {e_b_str}",
                    })

    clean = [r for r in reservations if id(r) not in to_remove]
    return {"warnings": warnings, "clean_reservations": clean}


DATE_COLS = {
    "Holiday Start Date": "%d/%m/%y",
    "Holiday End Date": "%d/%m/%y",
    "Departure Date": "%d/%m/%y",
    "Departure Flight Time": "%H:%M",
    "Flight Arrival Time": "%H:%M",
    "Flight Arrival Date": "%d/%m/%y",
}


# ── Weekly Report ───────────────────────────────────────────────────────────

def weekly_report(
    excel_path: str,
    caretakers_path: str,
    manual_reservations_path: str | None = None,
) -> str:
    """Generate caretaker-based weekly villa report.

    Args:
        excel_path: Path to the latest Resort Report .xlsx snapshot.
        caretakers_path: Path to caretakers.json assignments.
        manual_reservations_path: Optional path to manual_reservations.json.
            Manual reservations are merged with Excel data so they appear
            in the caretaker report.
    """
    caretakers = get_caretakers(caretakers_path)
    if caretakers and isinstance(caretakers[0], str):
        raise ValueError(f"Wrong file — expected caretakers.json, got a string list.")

    reservations = extract_reservations(excel_path)

    # Merge manual reservations so they appear in the weekly report
    manual_dicts = _load_manual_reservations_as_dicts(manual_reservations_path)
    if manual_dicts:
        reservations = list(reservations) + manual_dicts

    # Detect & deduplicate overlapping reservations
    overlap_result = _find_overlaps(reservations)
    reservations = overlap_result["clean_reservations"]
    overlap_warnings = overlap_result["warnings"]

    # Filter out past reservations: only show stays whose check-out is
    # today or later (current stays + upcoming). A stay that already
    # checked out is irrelevant to a caretaker's upcoming-week report.
    today = date.today()
    reservations = [
        r for r in reservations
        if not _is_past_reservation(r, today)
    ]

    reservations = sorted(
        reservations,
        key=lambda r: (
            pd.to_datetime(r.get("Holiday Start Date") or "", format="%d/%m/%y", errors="coerce"),
            str(r.get("Accomodation Name", "")),
        ),
    )
    reservations_by_villas = categorise_by_villas(reservations)

    for ct in caretakers:
        output = f"---- {ct['name']} ----\n"
        for assignment_name, extras_cfg in ct["assignments"].items():
            output += f"\n*{assignment_name}*\n"
            villa_name = f"Villa {assignment_name}"
            villa_res = reservations_by_villas.get(villa_name, [])
            if not villa_res:
                output += "\nReservasyon yok"
            else:
                for r in villa_res:
                    start = r.get("Holiday Start Date") or r.get("Holiday Start Day")
                    end = r.get("Holiday End Date")
                    output += f"{start} - {end}"
                    extra = ""
                    extras_text = r.get("ExtrasAggregated", "")
                    if extras_cfg.get("poolHeating") and "Pool Heating" in extras_text:
                        extra += " (Havuz ısıtması"
                    if extras_cfg.get("complimentaryCot") and "Complimentary Cot" in extras_text:
                        extra += ", bebek yatağı" if extra else " (Bebek yatağı"
                    if extras_cfg.get("welcomePack") and "Welcome Pack" in extras_text:
                        wp = extract_welcome_pack_size(extras_text)
                        extra += f", {wp})\n" if extra else f" ({wp})\n"
                    else:
                        extra += ")\n" if extra else "\n"
                    output += extra
        ct["output"] = output

    body = "\n\n".join([ct["output"] for ct in caretakers])

    # Prepend overlap / dedup warnings if any
    if overlap_warnings:
        warn_lines = ["⚠ UYARILAR", "─────────"]
        seen = set()  # deduplicate warning messages
        for w in overlap_warnings:
            villa_short = w["villa"].replace("Villa ", "")
            if w["type"] == "duplicate_removed":
                msg = (
                    f"🔁 {villa_short}: {w['removed']} manuel kaydı kaldırıldı "
                    f"— Excel'de {w['kept']} zaten var"
                )
            else:
                msg = (
                    f"⚠ {villa_short}: {w['range1']} ile {w['range2']} "
                    f"tarihleri çakışıyor!"
                )
            if msg not in seen:
                seen.add(msg)
                warn_lines.append(msg)
        warn_lines.append("")  # blank line before body
        return "\n".join(warn_lines) + "\n" + body

    return body


# ── İsmet Abi Report ────────────────────────────────────────────────────────

def ismet_abi_report(folder: str = "inputs/all_reservations") -> pd.DataFrame:
    """Reservations with Welcome Pack and/or Pool Heating extras."""
    files = list_files_with_extension(".xlsx", folder)
    all_res = []
    for file in files:
        all_res.append(pd.read_excel(file))
    if not all_res:
        return pd.DataFrame()
    merged = pd.concat(all_res, ignore_index=True)

    if "Opportunity Name" in merged.columns:
        def _has_relevant_extras(extras):
            if pd.isna(extras):
                return False
            return "Welcome Pack" in str(extras) or "Pool Heating" in str(extras)

        merged["__has"] = merged["ExtrasAggregated"].apply(_has_relevant_extras)
        merged = merged.sort_values(by=["Opportunity Name", "__has"], ascending=[True, False])
        merged = merged.drop_duplicates(subset=["Opportunity Name"], keep="first")
        merged = merged.drop(columns=["__has"])

    if "ExtrasAggregated" in merged.columns:
        merged["ExtrasAggregated"] = merged["ExtrasAggregated"].apply(_filter_extras_text)
        merged = merged[merged["ExtrasAggregated"].str.strip() != ""]

    if "Holiday Start Date" in merged.columns:
        merged["__sort"] = pd.to_datetime(merged["Holiday Start Date"], errors="coerce")
        sort_cols = ["__sort", "Accomodation Name"] if "Accomodation Name" in merged.columns else ["__sort"]
        merged = merged.sort_values(by=sort_cols, ascending=True)
        merged = merged.drop(columns=["__sort"])

    _format_date_columns(merged, DATE_COLS)
    return merged


# ── Hamdi Abi Report ────────────────────────────────────────────────────────

def hamdi_abi_report(inputs_folder: str = "inputs", caretakers_path: str = "inputs/caretakers.json") -> str:
    """Report for villas NOT assigned to any caretaker."""
    files = list_files_with_extension(".xlsx", inputs_folder)
    if not files:
        return "Excel dosyası bulunamadı"

    # Use the latest snapshot (by filename date), not an arbitrary first file
    latest = select_latest_file_with_extension(".xlsx", inputs_folder) or files[0]
    df = pd.read_excel(latest)
    assigned = set()
    caretakers = get_caretakers(caretakers_path)
    for ct in caretakers:
        assigned.update(f"Villa {v}" for v in ct["assignments"].keys())

    hamdi = df[~df["Accomodation Name"].isin(assigned)].copy()
    hamdi = hamdi[~hamdi["Accomodation Name"].str.contains("Total", na=False)]

    if "ExtrasAggregated" in hamdi.columns:
        hamdi["ExtrasAggregated"] = hamdi["ExtrasAggregated"].apply(_filter_extras_text)

    if "Holiday Start Date" in hamdi.columns:
        hamdi["__sort"] = pd.to_datetime(hamdi["Holiday Start Date"], errors="coerce")
        sort_cols = ["__sort", "Accomodation Name"] if "Accomodation Name" in hamdi.columns else ["__sort"]
        hamdi = hamdi.sort_values(by=sort_cols, ascending=True)
        hamdi = hamdi.drop(columns=["__sort"])

    _format_date_columns(hamdi, DATE_COLS)

    output = "---- Hamdi Abi ----\n"
    output += f"Kullanılan dosya: {os.path.basename(latest)}\n\n"

    if len(hamdi) == 0:
        output += "Rezervasyon yok"
    else:
        for villa_name, group in hamdi.groupby("Accomodation Name"):
            clean = villa_name.replace("Villa ", "")
            output += f"\n*{clean}*\n"
            for _, r in group.iterrows():
                start = r.get("Holiday Start Date", "")
                end = r.get("Holiday End Date", "")
                extras = r.get("ExtrasAggregated", "")
                if pd.notna(extras) and extras.strip():
                    output += f"{start} - {end} ({extras})\n"
                else:
                    output += f"{start} - {end}\n"

    # Stats
    all_villas = {v for v in df["Accomodation Name"].unique() if "Total" not in str(v)}
    hamdi_set = set(hamdi["Accomodation Name"].unique())
    output += f"\n\n--- İstatistik ---\n"
    output += f"Toplam: {len(all_villas)} | Atanmış: {len(assigned)} | Hamdi: {len(hamdi_set)}\n"
    return output


# ── Total History Report ────────────────────────────────────────────────────

def total_history_report(folder: str = "inputs/all_reservations") -> pd.DataFrame:
    """All reservations merged from every snapshot, deduplicated."""
    files = list_files_with_extension(".xlsx", folder)
    all_res = []
    for file in files:
        all_res.append(pd.read_excel(file))
    if not all_res:
        return pd.DataFrame()
    merged = pd.concat(all_res, ignore_index=True)

    if "Opportunity Name" in merged.columns:
        merged = merged.drop_duplicates(subset=["Opportunity Name"], keep="last")

    if "Holiday Start Date" in merged.columns:
        merged["__sort"] = pd.to_datetime(merged["Holiday Start Date"], errors="coerce")
        sort_cols = ["__sort", "Accomodation Name"] if "Accomodation Name" in merged.columns else ["__sort"]
        merged = merged.sort_values(by=sort_cols, ascending=True)
        merged = merged.drop(columns=["__sort"])

    _format_date_columns(merged, DATE_COLS)
    return merged


# ── Korsan Villas Report ────────────────────────────────────────────────────

def korsan_villas_report(
    excel_path: str,
    korsan_villas_json: str,
    template_path: str,
    year: int = 2026,
) -> openpyxl.Workbook:
    """Calendar-style Excel workbook — one sheet per month, May-Nov."""
    months = {
        5: ("Mayıs", 31), 6: ("Haziran", 30), 7: ("Temmuz", 31),
        8: ("Ağustos", 31), 9: ("Eylül", 30), 10: ("Ekim", 31), 11: ("Kasım", 30),
    }

    with open(korsan_villas_json, "r", encoding="utf-8") as f:
        villa_names = json.load(f)
    all_villas = {f"Villa {n}" for n in villa_names}

    df = pd.read_excel(excel_path)
    df.columns = [col.strip() for col in df.columns]
    df = df[df["Accomodation Name"].isin(all_villas)].copy()
    df = df[~df["Accomodation Name"].str.contains("Total", na=False)]

    if not pd.api.types.is_datetime64_any_dtype(df["Holiday Start Date"]):
        df["Holiday Start Date"] = pd.to_datetime(df["Holiday Start Date"], errors="coerce")
    if not pd.api.types.is_datetime64_any_dtype(df["Holiday End Date"]):
        df["Holiday End Date"] = pd.to_datetime(df["Holiday End Date"], errors="coerce")

    if "ExtrasAggregated" in df.columns:
        df["ExtrasAggregated"] = df["ExtrasAggregated"].apply(_filter_extras_text)

    wb = openpyxl.load_workbook(template_path)
    sorted_villas = sorted(all_villas, key=lambda v: v.replace("Villa ", "").lower())

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    villa_font = Font(name="Arial", size=11, bold=True)
    cell_font = Font(name="Arial", size=9)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    villa_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for month_num, (sheet_name, days_in_month) in months.items():
        ws = wb[sheet_name]
        month_start = datetime(year, month_num, 1)
        month_end = datetime(year, month_num, days_in_month, 23, 59, 59)
        ws.row_dimensions[1].height = 30

        villa_row_map = {}
        for i, vf in enumerate(sorted_villas):
            row = i + 2
            ws.row_dimensions[row].height = 30
            cell = ws.cell(row=row, column=1, value=vf.replace("Villa ", ""))
            cell.font = villa_font
            cell.alignment = villa_align
            villa_row_map[vf] = row

        max_name = max(len(v.replace("Villa ", "")) for v in sorted_villas)
        ws.column_dimensions["A"].width = max_name + 3
        for col_idx in range(2, days_in_month + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 4

        for _, res in df.iterrows():
            start, end = res["Holiday Start Date"], res["Holiday End Date"]
            if pd.isna(start) or pd.isna(end):
                continue
            villa = res["Accomodation Name"]
            if villa not in villa_row_map:
                continue
            res_start = max(start, month_start)
            res_end = min(end - timedelta(days=1), month_end)
            if res_start > res_end:
                continue

            passenger = str(res.get("Lead Passenger", "")) if not pd.isna(res.get("Lead Passenger")) else ""
            extras = str(res.get("ExtrasAggregated", "")) if not pd.isna(res.get("ExtrasAggregated")) else ""
            cell_text = f"{passenger}\n({extras})" if extras else passenger

            row = villa_row_map[villa]
            start_col, end_col = res_start.day + 1, res_end.day + 1

            def _apply(c, fill, font, align):
                c.fill, c.font, c.alignment = fill, font, align

            if start_col == end_col:
                c = ws.cell(row=row, column=start_col)
                c.value = cell_text
                _apply(c, green_fill, cell_font, cell_align)
                c.border = thin_border
            else:
                ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                c = ws.cell(row=row, column=start_col)
                c.value = cell_text
                _apply(c, green_fill, cell_font, cell_align)
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).fill = green_fill
                    ws.cell(row=row, column=col).border = thin_border

    return wb


# ── New Reservations Report ──────────────────────────────────────────────────

def new_reservations_report(
    folder: str,
    newer_path: str | None = None,
    manual_reservations_path: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> str:
    """Find reservations new in the latest (or specified) snapshot.

    Baseline = ALL previous snapshots + ALL manual reservations combined.
    New = reservations in the target snapshot whose Opportunity Name is NOT
          in the baseline.

    Matching is by Opportunity Name (unique booking ID).
    Optional date_from/date_to filter by Holiday Start Date (dd/mm/yy format).
    """
    import os as _os

    # ── Identify the "new" snapshot ──────────────────────────────────────
    all_files = _sorted_snapshots(folder)
    if not all_files:
        return "Snapshot dosyası bulunamadı."

    if newer_path:
        # User specified a particular file — use it as the "new" target
        target_file = newer_path
        # All OTHER files become the baseline
        baseline_files = [
            _os.path.join(folder, f)
            for f in all_files
            if _os.path.join(folder, f) != newer_path
        ]
    else:
        # Default: latest file is "new", everything else is baseline
        target_file = _os.path.join(folder, all_files[-1])
        baseline_files = [
            _os.path.join(folder, f) for f in all_files[:-1]
        ]
        if not baseline_files:
            return "Karşılaştırma için en az 2 snapshot dosyası gerekiyor."

    # ── Build baseline Opportunity Name set ──────────────────────────────
    baseline_ids: set[int] = set()

    # Load all previous snapshots
    for bf in baseline_files:
        try:
            dfb = pd.read_excel(bf)
            dfb.columns = [col.strip() for col in dfb.columns]
            if "Opportunity Name" in dfb.columns:
                for oid in dfb["Opportunity Name"].dropna():
                    try:
                        baseline_ids.add(int(oid))
                    except (ValueError, TypeError):
                        pass
        except Exception:
            continue  # skip corrupted/unreadable files

    # Merge manual reservations into baseline
    manual_dicts = _load_manual_reservations_as_dicts(manual_reservations_path)
    for m in manual_dicts:
        # Manual entries don't have Opportunity Name, but they represent
        # known bookings. We mark them via a synthetic ID that won't match
        # any real Opportunity Name. Instead, we compare by villa + date
        # overlap: if a reservation in the target snapshot overlaps a manual
        # entry for the same villa, it's NOT new.
        pass  # handled below via overlap check

    # ── Load the target snapshot ─────────────────────────────────────────
    try:
        df_new = pd.read_excel(target_file)
    except Exception:
        return f"Dosya okunamadı: {target_file}"

    df_new.columns = [col.strip() for col in df_new.columns]

    if "Opportunity Name" not in df_new.columns:
        return "Sütun bulunamadı: Opportunity Name"

    # Exclude "Total" summary rows early
    if "Accomodation Name" in df_new.columns:
        df_new = df_new[~df_new["Accomodation Name"].str.contains("Total", na=False)]

    # Find new Opportunity Names (not in baseline snapshots)
    new_opp = df_new["Opportunity Name"].fillna(-1).astype(int).values
    new_mask = [oid not in baseline_ids and oid != -1 for oid in new_opp]
    df_new_res = df_new[new_mask].copy()

    # Also remove entries that EXACTLY match a manual reservation
    # (same villa + same start + same end → pre-announced, not new)
    if manual_dicts and len(df_new_res) > 0:
        manual_by_villa: dict[str, list[dict]] = {}
        for m in manual_dicts:
            v = m.get("Accomodation Name", "")
            if v:
                manual_by_villa.setdefault(v, []).append(m)

        if manual_by_villa:
            drop_indices: list[int] = []
            for idx, row in df_new_res.iterrows():
                villa = str(row.get("Accomodation Name", ""))
                if villa not in manual_by_villa:
                    continue
                row_start = _to_date(
                    row.get("Holiday Start Date") or row.get("Holiday Start Day")
                )
                row_end = _to_date(row.get("Holiday End Date"))
                if row_start is None or row_end is None:
                    continue
                for m in manual_by_villa.get(villa, []):
                    m_start = _parse_date_dmy(m.get("Holiday Start Date", ""))
                    m_end = _parse_date_dmy(m.get("Holiday End Date", ""))
                    if m_start is None or m_end is None:
                        continue
                    # Exact same dates → same reservation, already known
                    if row_start == m_start and row_end == m_end:
                        drop_indices.append(idx)
                        break
            if drop_indices:
                df_new_res = df_new_res.drop(index=drop_indices)

    # ── Parse dates for filtering & output ───────────────────────────────
    if "Holiday Start Date" in df_new_res.columns:
        if not pd.api.types.is_datetime64_any_dtype(df_new_res["Holiday Start Date"]):
            df_new_res["Holiday Start Date"] = pd.to_datetime(
                df_new_res["Holiday Start Date"], errors="coerce"
            )

        # Apply date range filter
        if date_from:
            df_from = pd.to_datetime(date_from, format="%d/%m/%y", errors="coerce")
            if pd.notna(df_from):
                df_new_res = df_new_res[df_new_res["Holiday Start Date"] >= df_from]
        if date_to:
            df_to = pd.to_datetime(date_to, format="%d/%m/%y", errors="coerce")
            if pd.notna(df_to):
                df_new_res = df_new_res[df_new_res["Holiday Start Date"] <= df_to]

        # Format dates for output
        df_new_res["Holiday Start Date"] = df_new_res["Holiday Start Date"].dt.strftime("%d/%m/%y")

    if "Holiday End Date" in df_new_res.columns:
        if not pd.api.types.is_datetime64_any_dtype(df_new_res["Holiday End Date"]):
            df_new_res["Holiday End Date"] = pd.to_datetime(
                df_new_res["Holiday End Date"], errors="coerce"
            )
        df_new_res["Holiday End Date"] = df_new_res["Holiday End Date"].dt.strftime("%d/%m/%y")

    # ── Sort ─────────────────────────────────────────────────────────────
    sort_cols = []
    if "Accomodation Name" in df_new_res.columns:
        sort_cols.append("Accomodation Name")
    if "Holiday Start Date" in df_new_res.columns:
        df_new_res["__sort_date"] = pd.to_datetime(
            df_new_res["Holiday Start Date"], format="%d/%m/%y", errors="coerce"
        )
        sort_cols.append("__sort_date")

    if sort_cols:
        df_new_res = df_new_res.sort_values(by=sort_cols, ascending=True)

    # ── Build output ─────────────────────────────────────────────────────
    lines = ["YENİ REZERVASYONLAR"]
    snap_label = _os.path.basename(target_file)
    lines.append(f"Hedef snapshot: {snap_label}")
    lines.append(f"Baseline: {len(baseline_files)} önceki snapshot"
                 + (f" + {len(manual_dicts)} manuel kayıt" if manual_dicts else ""))

    if date_from or date_to:
        range_str = f"{date_from or '…'} - {date_to or '…'}"
        lines.append(f"Tarih aralığı: {range_str}")

    lines.append(f"Toplam: {len(df_new_res)} yeni rezervasyon")
    lines.append("")

    if len(df_new_res) == 0:
        lines.append("Yeni rezervasyon bulunamadı.")
        return "\n".join(lines)

    # Group by villa
    if "Accomodation Name" in df_new_res.columns:
        current_villa = None
        for _, r in df_new_res.iterrows():
            villa = str(r.get("Accomodation Name", ""))
            if villa != current_villa:
                current_villa = villa
                clean = villa.replace("Villa ", "")
                lines.append(clean)

            start = r.get("Holiday Start Date", "")
            end = r.get("Holiday End Date", "")
            pax = str(r.get("Lead Passenger", "")) if pd.notna(r.get("Lead Passenger")) else ""
            extras = str(r.get("ExtrasAggregated", "")) if pd.notna(r.get("ExtrasAggregated")) else ""

            line = f"{start} - {end}"
            if pax:
                line += f"  {pax}"
            extras_filtered = _filter_extras_text(extras)
            if extras_filtered:
                line += f"  ({extras_filtered})"

            lines.append(line)
        lines.append("")
    else:
        for _, r in df_new_res.iterrows():
            lines.append(str(r.to_dict()))

    return "\n".join(lines)


# ── Change Detection Report ──────────────────────────────────────────────────

def change_detection_report(
    folder: str,
    target_path: str | None = None,
    manual_reservations_path: str | None = None,
) -> dict:
    """Detect new, deleted, and conflicting reservations between snapshots.

    Three pools:
      Pool 1 (baseline) = all snapshots EXCEPT target + all manual reservations
      Pool 3 (target)   = the target snapshot (default: latest)

    Returns a structured dict:
      - new: reservations in target not in baseline
      - deleted: reservations in baseline not in target whose dates suggest
                 they should still be present (cancelled)
      - conflicts: overlapping reservations within the target snapshot
      - summary: counts
    """
    import os as _os

    all_files = _sorted_snapshots(folder)
    if not all_files:
        return {"error": "Snapshot dosyası bulunamadı."}

    # ── Determine target snapshot ────────────────────────────────────────
    if target_path:
        target_file = target_path
        baseline_files = [
            _os.path.join(folder, f)
            for f in all_files
            if _os.path.join(folder, f) != target_path
        ]
    else:
        target_file = _os.path.join(folder, all_files[-1])
        baseline_files = [_os.path.join(folder, f) for f in all_files[:-1]]

    target_date = _extract_snapshot_date(_os.path.basename(target_file))

    # ── Load target ──────────────────────────────────────────────────────
    try:
        df_t = pd.read_excel(target_file)
    except Exception:
        return {"error": f"Dosya okunamadı: {target_file}"}
    df_t.columns = [col.strip() for col in df_t.columns]
    if "Accomodation Name" in df_t.columns:
        df_t = df_t[~df_t["Accomodation Name"].str.contains("Total", na=False)]

    manual_dicts = _load_manual_reservations_as_dicts(manual_reservations_path)

    # ── Build baseline (all previous snapshots) ──────────────────────────
    # baseline[opp_name] = {villa, start, end, passenger, last_seen}
    baseline: dict[int, dict] = {}
    if "Opportunity Name" in df_t.columns:
        for bf in baseline_files:
            try:
                dfb = pd.read_excel(bf)
            except Exception:
                continue
            dfb.columns = [col.strip() for col in dfb.columns]
            if "Opportunity Name" not in dfb.columns:
                continue
            if "Accomodation Name" in dfb.columns:
                dfb = dfb[~dfb["Accomodation Name"].str.contains("Total", na=False)]
            for _, r in dfb.iterrows():
                try:
                    opp = int(r["Opportunity Name"])
                except (ValueError, TypeError):
                    continue
                # Keep the earliest start date seen (most complete info)
                start = _to_date(r.get("Holiday Start Date"))
                if opp in baseline:
                    # keep earliest start (booking start doesn't change)
                    if start and (baseline[opp]["start"] is None or start < baseline[opp]["start"]):
                        baseline[opp]["start"] = start
                    # last_seen = most recent file
                    baseline[opp]["last_seen"] = _os.path.basename(bf)
                    continue
                baseline[opp] = {
                    "villa": str(r.get("Accomodation Name", "") or ""),
                    "start": start,
                    "end": _to_date(r.get("Holiday End Date")),
                    "passenger": str(r.get("Lead Passenger", "") or "") if pd.notna(r.get("Lead Passenger")) else "",
                    "last_seen": _os.path.basename(bf),
                }

    # ── Target opportunity set ───────────────────────────────────────────
    target_opps: set[int] = set()
    if "Opportunity Name" in df_t.columns:
        for oid in df_t["Opportunity Name"].dropna():
            try:
                target_opps.add(int(oid))
            except (ValueError, TypeError):
                pass

    # ── NEW: in target, not in baseline ──────────────────────────────────
    # Build manual exact-match set: (villa, start, end) tuples that mean
    # the Excel row was pre-announced and is NOT new.
    manual_exact: set[tuple] = set()
    for m in manual_dicts:
        ms = _parse_date_dmy(m.get("Holiday Start Date", ""))
        me = _parse_date_dmy(m.get("Holiday End Date", ""))
        if ms is not None and me is not None:
            manual_exact.add((m.get("Accomodation Name", ""), ms, me))

    new_list = []
    if "Opportunity Name" in df_t.columns:
        for _, r in df_t.iterrows():
            try:
                opp = int(r["Opportunity Name"])
            except (ValueError, TypeError):
                continue
            if opp in baseline:
                continue
            villa = str(r.get("Accomodation Name", "") or "")
            rs = _to_date(r.get("Holiday Start Date"))
            re_ = _to_date(r.get("Holiday End Date"))
            # Skip if a manual reservation has the exact same villa+dates
            if rs is not None and re_ is not None and (villa, rs, re_) in manual_exact:
                continue
            new_list.append({
                "villa": villa,
                "start": _fmt_date(r.get("Holiday Start Date")),
                "end": _fmt_date(r.get("Holiday End Date")),
                "passenger": str(r.get("Lead Passenger", "") or "") if pd.notna(r.get("Lead Passenger")) else "",
                "opportunity": opp,
            })

    # ── DELETED: in the SECOND-TO-LATEST snapshot, not in target ─────────
    # Only compare the previous snapshot (not all history) so we surface
    # only the LATEST cancellations.
    # A snapshot dated X contains only reservations with start >= X.
    deleted_list = []
    prev_baseline: dict[int, dict] = {}
    if len(all_files) >= 2:
        prev_file = _os.path.join(folder, all_files[-2])
        try:
            dfp = pd.read_excel(prev_file)
            dfp.columns = [col.strip() for col in dfp.columns]
            if "Accomodation Name" in dfp.columns:
                dfp = dfp[~dfp["Accomodation Name"].str.contains("Total", na=False)]
            if "Opportunity Name" in dfp.columns:
                for _, r in dfp.iterrows():
                    try:
                        opp = int(r["Opportunity Name"])
                    except (ValueError, TypeError):
                        continue
                    prev_baseline[opp] = {
                        "villa": str(r.get("Accomodation Name", "") or ""),
                        "start": _to_date(r.get("Holiday Start Date")),
                        "end": _to_date(r.get("Holiday End Date")),
                        "passenger": str(r.get("Lead Passenger", "") or "") if pd.notna(r.get("Lead Passenger")) else "",
                        "last_seen": _os.path.basename(prev_file),
                    }
        except Exception:
            pass

    if target_date:
        for opp, info in prev_baseline.items():
            if opp in target_opps:
                continue
            start = info.get("start")
            # Expected in target if it starts on/after the snapshot date
            if start is not None and start >= datetime.combine(target_date, datetime.min.time()):
                deleted_list.append({
                    "villa": info["villa"],
                    "start": _fmt_date(start),
                    "end": _fmt_date(info.get("end")),
                    "passenger": info.get("passenger", ""),
                    "opportunity": opp,
                    "last_seen_in": info.get("last_seen", ""),
                })

    # ── CONFLICTS: overlapping reservations within the target ────────────
    conflict_list = []
    if "Opportunity Name" in df_t.columns:
        target_rows = [
            {
                "Accomodation Name": str(r.get("Accomodation Name", "") or ""),
                "Holiday Start Date": _fmt_date(r.get("Holiday Start Date")),
                "Holiday End Date": _fmt_date(r.get("Holiday End Date")),
            }
            for _, r in df_t.iterrows()
        ]
        overlap_result = _find_overlaps(target_rows)
        for w in overlap_result["warnings"]:
            if w["type"] == "overlap_warning":
                conflict_list.append({
                    "villa": w["villa"],
                    "range1": w["range1"],
                    "range2": w["range2"],
                })

    # ── Summary ──────────────────────────────────────────────────────────
    summary = {
        "new_count": len(new_list),
        "deleted_count": len(deleted_list),
        "conflict_count": len(conflict_list),
    }

    return {
        "target_file": _os.path.basename(target_file),
        "target_date": target_date.isoformat() if target_date else None,
        "baseline_snapshot_count": len(baseline_files),
        "manual_count": len(manual_dicts),
        "new": new_list,
        "deleted": deleted_list,
        "conflicts": conflict_list,
        "summary": summary,
    }


def _fmt_date(val) -> str:
    """Format a date value (Timestamp/datetime/str) as dd/mm/yy string."""
    dt = _to_date(val)
    if dt is None:
        return str(val) if val is not None else ""
    return dt.strftime("%d/%m/%y")


# ── Korsan Check-ins Report ─────────────────────────────────────────────────

def korsan_checkins_report(
    korsan_villas_json: str,
    timelines: dict,
    from_date: date | None = None,
    days: int = 10,
) -> str:
    """List upcoming check-ins (arrival dates) for each Korsan villa.

    Window = [from_date, from_date + days], where from_date defaults to
    today - 1 day (yesterday). For each Korsan villa, lists every reservation
    whose start_date (check-in) falls in the window.
    """
    if from_date is None:
        from_date = date.today() - timedelta(days=1)

    end_date = from_date + timedelta(days=days)

    # Load Korsan villa short names
    try:
        with open(korsan_villas_json, "r", encoding="utf-8") as f:
            villa_names = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return "Korsan villa listesi bulunamadı."

    lines = ["KORSAN CHECK-IN'LER"]
    lines.append(f"{from_date.strftime('%d/%m/%y')} - {end_date.strftime('%d/%m/%y')}")
    lines.append("")

    for short in villa_names:
        # Timeline keys may be short ("Au Soleil") or prefixed ("Villa Au Soleil")
        timeline = timelines.get(short) or timelines.get(f"Villa {short}")

        # Collect check-ins (start_date) within the window
        checkins = []
        if timeline is not None:
            for r in timeline.records:
                if from_date <= r.start_date <= end_date:
                    checkins.append(r)

        checkins.sort(key=lambda r: r.start_date)

        lines.append(short)
        if not checkins:
            lines.append("  (check-in yok)")
        else:
            for c in checkins:
                pax = f"  {c.lead_passenger}" if c.lead_passenger else ""
                lines.append(f"  {c.start_date.strftime('%d/%m/%y')}{pax}")
        lines.append("")

    return "\n".join(lines)
